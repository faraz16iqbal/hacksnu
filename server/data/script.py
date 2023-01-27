import openpyxl
import json
# load excel with its path
wrkbk = openpyxl.load_workbook("dataset.xlsx")
sh = wrkbk.active

data = []

s1 = set()
s2 = set()
data = {}
d = []
for i in range(2, 10):      
    cur = {}
    if(len(d) == 5):
        d.clear()
        s1.clear()
        continue
    for j in range(1, sh.max_column+1):
        cell_obj = sh.cell(row=i, column=j)
        if j == 1:
            cur["country"] = cell_obj.value
        elif j == 2:
            cur["items"] = cell_obj.value
            if(cell_obj.value in s1):
                continue
            s1.add(cell_obj.value)
        if j == 5:
            cur["type"] = cell_obj.value
    d.append(cur)    
    if(cur["country"] not in data.keys()):
        data[cur["country"]] = []
        data[cur["country"]].append(d)
    else:
        data[cur["country"]].append(d)

# print(len(data))
print(data)
            
        
